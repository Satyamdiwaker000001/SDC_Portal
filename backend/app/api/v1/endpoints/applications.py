from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
import csv
import io
import os
import shutil
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
from ....api import deps
from ....models.models import Application, User, Setting
from ....schemas.application import ApplicationCreate, ApplicationOut

router = APIRouter()

@router.post("/", status_code=status.HTTP_201_CREATED)
def submit_enrollment(
    *,
    db: Session = Depends(deps.get_db),
    application_in: ApplicationCreate,
) -> Any:
    """
    Public recruitment enrollment.
    """
    application = Application(
        name=application_in.name,
        email=application_in.email,
        contact=application_in.contact,
        class_name=application_in.class_name,
        interested=application_in.interested,
        photo_url=application_in.photo_url,
        resume_url=application_in.resume_url
    )
    db.add(application)
    db.commit()
    db.refresh(application)
    return {"status": "ENROLLMENT_RECEIVED", "id": application.id}

@router.post("/submit", status_code=status.HTTP_201_CREATED)
async def submit_multimedia_enrollment(
    *,
    db: Session = Depends(deps.get_db),
    name: str = Form(...),
    email: str = Form(...),
    contact: str = Form(...),
    class_name: str = Form(...),
    interested: str = Form(...),
    photo: Optional[UploadFile] = File(None),
    resume: Optional[UploadFile] = File(None),
) -> Any:
    """
    Public recruitment enrollment with file support [PHOTO & RESUME].
    """
    # Setup storage [TACTICAL_STORAGE]
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    static_dir = os.path.join(base_dir, "static")
    
    photo_url = None
    resume_url = None

    if photo:
        p_ext = photo.filename.split(".")[-1]
        p_name = f"photo_{int(datetime.utcnow().timestamp())}_{name.replace(' ', '_')}.{p_ext}"
        p_path = os.path.join(static_dir, "applicants", p_name)
        os.makedirs(os.path.dirname(p_path), exist_ok=True)
        with open(p_path, "wb") as buffer:
            shutil.copyfileobj(photo.file, buffer)
        photo_url = f"http://127.0.0.1:8000/static/applicants/{p_name}"

    if resume:
        r_ext = resume.filename.split(".")[-1]
        r_name = f"resume_{int(datetime.utcnow().timestamp())}_{name.replace(' ', '_')}.{r_ext}"
        r_path = os.path.join(static_dir, "resumes", r_name)
        os.makedirs(os.path.dirname(r_path), exist_ok=True)
        with open(r_path, "wb") as buffer:
            shutil.copyfileobj(resume.file, buffer)
        resume_url = f"http://127.0.0.1:8000/static/resumes/{r_name}"

    application = Application(
        name=name,
        email=email,
        contact=contact,
        class_name=class_name,
        interested=interested,
        photo_url=photo_url,
        resume_url=resume_url
    )
    db.add(application)
    db.commit()
    db.refresh(application)
    return {"status": "TRANS_COMPLETE", "id": application.id}

@router.get("/", response_model=List[ApplicationOut])
def read_aspirants(
    db: Session = Depends(deps.get_db),
    skip: int = 0,
    limit: int = 100,
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Retrieve aspirant registry (Admin only).
    """
    applicants = db.exec(select(Application).offset(skip).limit(limit)).all()
    return applicants

@router.get("/status")
def get_recruitment_status(
    db: Session = Depends(deps.get_db),
) -> Any:
    """
    Check if recruitment is live.
    """
    setting = db.get(Setting, "isRecruitmentOpen")
    if not setting:
        return {"isOpen": False}
    return {"isOpen": setting.value == "true"}

@router.post("/toggle")
def toggle_recruitment(
    db: Session = Depends(deps.get_db),
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Toggle recruitment status (Admin only).
    """
    setting = db.get(Setting, "isRecruitmentOpen")
    if not setting:
        # Default to true if it didn't exist
        new_status = True
        setting = Setting(key="isRecruitmentOpen", value="true")
    else:
        new_status = setting.value != "true"
        setting.value = str(new_status).lower()
    
    db.add(setting)
    db.commit()
    return {"status": "UPDATED", "isOpen": new_status}

@router.get("/export-csv")
def export_aspirants_csv(
    db: Session = Depends(deps.get_db),
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Export all aspirants as CSV (Admin only).
    """
    applicants = db.exec(select(Application)).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Email", "Contact", "Class", "Interests", "Status", "Photo_Link", "Resume_Link", "Timestamp"])
    
    for row in applicants:
        writer.writerow([
            row.id, row.name, row.email, row.contact, 
            row.class_name, row.interested, row.status,
            row.photo_url or "N/A", row.resume_url or "N/A",
            row.timestamp
        ])
    
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=aspirants_registry.csv"}
    )

@router.get("/export-excel")
def export_aspirants_excel(
    db: Session = Depends(deps.get_db),
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Generate high-fidelity Excel report with embedded photos.
    """
    applicants = db.exec(select(Application)).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Aspirants_Registry"
    
    # Headers [INTEL_COLUMNS]
    headers = ["ID", "Photo", "Name", "Email", "Contact", "Class", "Interests", "Resume_Link", "Status", "Timestamp"]
    ws.append(headers)
    
    # Adjust column widths
    ws.column_dimensions['B'].width = 15 # Photo column
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 25

    for i, row in enumerate(applicants, start=2):
        ws.row_dimensions[i].height = 60 # Set row height for photo
        
        # Write text data
        ws.cell(row=i, column=1, value=row.id)
        ws.cell(row=i, column=3, value=row.name)
        ws.cell(row=i, column=4, value=row.email)
        ws.cell(row=i, column=5, value=row.contact)
        ws.cell(row=i, column=6, value=row.class_name)
        ws.cell(row=i, column=7, value=row.interested)
        ws.cell(row=i, column=8, value=row.resume_url or "N/A")
        ws.cell(row=i, column=9, value=row.status)
        ws.cell(row=i, column=10, value=str(row.timestamp))
        
        # Embed Photo [LENS_CAPTURE]
        if row.photo_url and row.photo_url != "N/A":
            try:
                # In local dev, we might need to map URL to path if server isn't up
                # For this implementation, we fetch via requests for robustness
                response = requests.get(row.photo_url, timeout=5)
                if response.status_code == 200:
                    img_data = BytesIO(response.content)
                    img = OpenpyxlImage(img_data)
                    img.width = 60
                    img.height = 60
                    ws.add_image(img, f'B{i}')
            except Exception as e:
                print(f"FAILED_PHOTO_EMBED: {row.id} - {str(e)}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=recruitment_report.xlsx"}
    )

@router.delete("/{application_id}")
def discard_aspirant(
    *,
    db: Session = Depends(deps.get_db),
    application_id: int,
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Remove an aspirant record (Admin only).
    """
    app = db.get(Application, application_id)
    if not app:
        raise HTTPException(status_code=404, detail="Aspirant record not found")
    db.delete(app)
    db.commit()
    return {"status": "DISCARDED"}

@router.post("/{application_id}/approve")
def enlist_aspirant(
    *,
    db: Session = Depends(deps.get_db),
    application_id: int,
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Approve an aspirant and convert to Operative (Admin only).
    """
    app = db.get(Application, application_id)
    if not app:
        raise HTTPException(status_code=404, detail="Aspirant record not found")
    
    # Check if user already exists
    existing_user = db.exec(select(User).where(User.email == app.email)).first()
    if existing_user:
        db.delete(app) # Clean up application if they already have an account
        db.commit()
        return {"status": "ALREADY_ENLISTED"}
    
    # Create new operative record - [SCHEMA_ALIGNED]
    new_user = User(
        id=app.email, # Use email or a new ID protocol
        name=app.name,
        email=app.email,
        hashed_password="hashed_placeholder_123", # System-generated hash placeholder
        role="developer"
    )
    db.add(new_user)
    db.delete(app) # Remove from pipeline
    db.commit()
    return {"status": "RECRUIT_ENLISTED"}

@router.post("/bulk-approve")
def bulk_enlist_aspirants(
    *,
    db: Session = Depends(deps.get_db),
    ids: List[int],
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Approve multiple aspirants and convert to Operatives (Admin only).
    """
    processed = 0
    skipped = 0
    for app_id in ids:
        app = db.get(Application, app_id)
        if not app:
            skipped += 1
            continue
            
        existing_user = db.exec(select(User).where(User.email == app.email)).first()
        if existing_user:
            db.delete(app)
            processed += 1
            continue
            
        new_user = User(
            id=app.email,
            name=app.name,
            email=app.email,
            hashed_password="hashed_placeholder_123", # System-generated
            role="developer"
        )
        db.add(new_user)
        db.delete(app)
        processed += 1
    
    db.commit()
    return {"status": "BULK_ENLISTMENT_COMPLETE", "processed": processed, "skipped": skipped}

@router.post("/bulk-reject")
def bulk_discard_aspirants(
    *,
    db: Session = Depends(deps.get_db),
    ids: List[int],
    current_admin: User = Depends(deps.get_current_active_admin),
) -> Any:
    """
    Remove multiple aspirant records in bulk (Admin only).
    """
    discarded = 0
    for app_id in ids:
        app = db.get(Application, app_id)
        if app:
            db.delete(app)
            discarded += 1
            
    db.commit()
    return {"status": "BULK_DISCARD_COMPLETE", "discarded": discarded}
